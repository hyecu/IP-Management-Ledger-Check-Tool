from PyQt5.QtGui import QIcon, QFont, QDesktopServices,QColor
from PyQt5.QtCore import QUrl, QThread, pyqtSignal, Qt
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, NamedStyle, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from PyQt5.QtWidgets import *
from queue import Queue
from scapy.all import *
import threading
import psutil
import sys
import socket
import ipaddress
import re
import pandas as pd
import os
#import time

# class progress(QThread):
#     update_progress = pyqtSignal(int)
    
#     def run(self):
#         total = 100
#         for i in range(total + 1):
#             # 작업 수행
#             time.sleep(0.1)dmd

#             # 진행률 업데이트
#             self.update_progress.emit(i)

#         # 작업 완료 후 시그널 발생
#         self.finished.emit()

# 전처리 (엑셀 기준으로 데이터가 존재하는 행과 열부터 시작하도록 DataFrame 자르기)
def preprocessing(data):
    # NaN 값을 공백 문자열로 대체
    data.fillna('', inplace=True)
    data = data[data.apply(lambda row: any(row.notna()), axis=1)]
    if 'Unnamed' in data.columns[0]:
        data = data.rename(columns=data.iloc[0])
    data = data.dropna(axis='columns', how='all')
    data = data[1:]
    data.columns = [str(i) if pd.isna(col) else col for i, col in enumerate(data.columns, 1)]
    return data

# IP 주소 컬럼이름
def get_ip_columns(data, pattern):
    ip_columns = [column for column in data.columns if data[column].apply(lambda x: isinstance(x, str) and re.match(pattern, x)).any()]
    return ip_columns

# 호스트 컬럼이름
def get_host_column(data):
    for column in data.columns:
        if 'host' in column.lower() or '호스트' in column:
            host_column = column
            break
    return host_column

# 정상 IP 주소 필터링 + (비정상 IP 주소는 unnormal_ip_L에 추가, 리턴 안해도 반영됨)
def filtering_ip(data, pattern, ip_columns, unnormal_ip_L):
    dropped_data = data.copy()
    dropped_data.dropna(subset=[ip_columns[0]], inplace=True)
    all_ip_excel_L = dropped_data[ip_columns[0]].astype(str).to_list()
    normal_ip_DF = pd.DataFrame(columns=data.columns)

    for ip in all_ip_excel_L:
        if re.match(pattern, ip):
            normal_ip_DF = pd.concat([normal_ip_DF, data.loc[(data[ip_columns] == ip).all(axis=1), :]], ignore_index=True)
        else:
            unnormal_ip_L.append(ip)
    
    return normal_ip_DF

# 관리대장 IP, 스캔IP 비교 결과 분류(3가지 경우의 수)
def classification_data(excel_dict, unnormal_ip_L, scan_result, ip_columns, host_column, result):

    try:
        excel_ip_L = list(excel_dict.keys())
        scan_ip_L = list(scan_result.keys())
        excel_set = set(excel_ip_L)
        scan_set = set(scan_ip_L)

        matching_ip = excel_set.intersection(scan_set)
        missing_from_scan = excel_set - matching_ip
        missing_from_excel = scan_set - matching_ip

        matching_ip = list(matching_ip)
        missing_from_scan = list(missing_from_scan)
        missing_from_excel = list(missing_from_excel)
        
        chk_null_ip = result[ip_columns[0]].isna().sum()
        
        if matching_ip:
            result = process_matching_data(matching_ip, excel_dict, scan_result, ip_columns, result)
        if missing_from_scan:
            result = process_missing_scan(missing_from_scan, excel_dict,  ip_columns, host_column, result)
        if missing_from_excel:
            result = process_missing_excel(missing_from_excel, ip_columns, scan_result, host_column, result)
        if unnormal_ip_L:
            result = process_unnormal_ip(unnormal_ip_L, ip_columns, result)
        if chk_null_ip:
            result = process_null_ip(result, ip_columns)
    except:
        return 
    else:
        return result

# 대장o, 스캔o
def process_matching_data(matching_ip, excel_dict, scan_result, ip_columns, result):
    # 관리대장: ip랑 host 딕셔너리로 결합
    # ip만 따로 추출해서 3가지 경우로 분류
    # 일치하면? ip에 해당하는 딕셔너리(host)랑 스캔ip에 해당하는 딕셔너리(host) 비교
    for ip in matching_ip:
        excel_host = excel_dict[ip]
        scan_host = scan_result[ip]

        if excel_host != scan_host:
            result.loc[result[ip_columns[0]] == ip, 'result'] = '호스트 불일치'
    return result

# 대장o, 스캔x
def process_missing_scan(missing_from_scan, excel_dict, ip_columns, host_column, result):
    #대장에만 있는 ip를 주면 ↓출력
    #"대장o 스캔x
    # host: 대장host이름"
    for ip in missing_from_scan:
        excel_host = excel_dict[ip]
        result.loc[result[ip_columns[0]] == ip, 'result'] = f"대장o, 스캔x"
        result.loc[result[ip_columns[0]] == ip, host_column ] = f"{excel_host}"
        # result.loc[result[ip_columns[0]] == ip, 'result'] = f"대장o, 스캔x  (host: {excel_host})"
    return result

# 대장x, 스캔o (IP 검증 결과를 저장할 DataFrame 및 컬럼 만들기)
def process_missing_excel(missing_from_excel, ip_columns, scan_result, host_column, result):
    # 대장에 없는 IP 처리하기
    if missing_from_excel:
        # 기존 관리대장에 새로 추가
        new_rows = pd.DataFrame(columns=result.columns)
        new_rows[ip_columns[0]] = missing_from_excel
        #new_rows['result'] = new_rows[ip_columns[0]].map(scan_result).apply(lambda x: f'대장x 스캔o' if x is not None else f'대장x 스캔o (host: {new_rows.at[x, host_column] if x is not None else "nan"})')
        new_rows['result'] = new_rows[ip_columns[0]].map(scan_result).apply(lambda x: f'대장x 스캔o  (host: {x})' if x is not None else '대장x 스캔o  (host: nan)')
        result = pd.concat([result, new_rows], ignore_index=True)  #병합
    return result

# IP 주소 오류
def process_unnormal_ip(unnormal_ip_L, ip_columns, result):
    for unnormal_ip in unnormal_ip_L:
        result.loc[result[ip_columns[0]].astype(str) == str(unnormal_ip), 'result'] = 'Invalid IP Address, 스캔불가'
    return result

# IP 주소가 존재하지 않음
def process_null_ip(result, ip_columns):
    result.loc[result[ip_columns[0]].isna(), 'result'] = 'No IP Address, 스캔불가'
    return result

# 검증 수행
# def Verification(file, scan_result, output_file):
def Verification(file, scan_result):
    try:
        # IP 관리대장 파일 불러오기
        data = pd.read_excel(file)
        # 전처리 (엑셀 기준으로 데이터가 존재하는 행과 열부터 시작하도록 DataFrame 자르기)
        data = preprocessing(data)
        # IP 주소 정규표현식
        pattern = r'^(?!255\.255\.255\.(?:128|192|224|240|248|252|255)$)(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}(?!0)([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])$'
        # 결과 저장할 DataFrame
        result = data.copy()
        result['result'] = ''
        # IP 컬럼 가져오기
        ip_columns = get_ip_columns(data, pattern)
        # 정상 IP 주소 값만 필터링 + (비정상 IP 주소 저장할 배열 만들기)
        unnormal_ip_L = []
        normal_ip_DF = filtering_ip(data, pattern, ip_columns, unnormal_ip_L)
        
        # 호스트 컬럼 가져오기
        host_column = get_host_column(data)

        #관리대장 ip(key), host(value) 딕셔너리
        excel_dict = dict(zip(normal_ip_DF[ip_columns[0]], normal_ip_DF[host_column]))
        
        # 대장 처리 및 스캔 결과 처리
        result = classification_data(excel_dict, unnormal_ip_L, scan_result, ip_columns, host_column, result)
        result['#'] = ''
        result_all_data = result[['#',str(ip_columns[0]),host_column,'result']]
        result_all_data.index += 1
        result_all_data = result_all_data.reset_index()
        result_all_data.fillna('', inplace=True)
        
        # 결과 저장
        # save_to_excel(result, ip_columns, host_column, output_file)
    except Exception as error:
        print(error)
        return False
    else:
        return result_all_data


# 엑셀 출력
def save_to_excel(result, ip_column, host_column, output_file):
    # 결과 엑셀 파일(.xlsx)로 출력하기
    workbook = Workbook()
    worksheet = workbook.active

    # 데이터프레임을 엑셀에 저장
    for r in dataframe_to_rows(result, index=False, header=True):
        worksheet.append(r)

    # 데이터 있는 곳에 모든 테두리 적용
    apply_borders(worksheet, start_row=1, start_col=1, end_row=worksheet.max_row, end_col=worksheet.max_column)
    
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            if cell.column == result.columns.get_loc('result') + 1:
                if '대장x' in cell.value:
                    # 대장x인 행에 대해서만 빨간색 적용
                    for column in worksheet.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=worksheet.max_column):
                        for cell in column:
                            apply_color(worksheet, row=cell.row, column=cell.column, color='F6CECE')

                    # result 컬럼에 색상 적용 (더 진한 빨간색)
                    apply_color(worksheet, row=cell.row, column=cell.column, color='FF0000')

                    # IP 주소 컬럼에 색상 적용 (더 진한 빨간색)
                    ip_column_index = result.columns.get_loc(ip_column[0]) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='FF0000')
                    
                elif '스캔x' in cell.value:
                    # 스캔x인 행에 대해서만 노란색 적용
                    for column in worksheet.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=worksheet.max_column):
                        for cell in column:
                            apply_color(worksheet, row=cell.row, column=cell.column, color='F6F5CC')

                    # result 컬럼에 색상 적용 (더 진한 노란색)
                    apply_color(worksheet, row=cell.row, column=cell.column, color='FFFF00')

                    # 호스트 컬럼에 색상 적용 (더 진한 노란색)
                    ip_column_index = result.columns.get_loc(host_column) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='FFFF00')

                    # IP 주소 컬럼에 색상 적용 (더 진한 노란색)
                    ip_column_index = result.columns.get_loc(ip_column[0]) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='FFFF00')
                    
                elif '호스트 불일치' == cell.value:
                    # '호스트 불일치' 행에 대해서만 색상 적용(실버블론드)
                    for column in worksheet.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=worksheet.max_column):
                        for cell in column:
                            apply_color(worksheet, row=cell.row, column=cell.column, color='FFFFCC')

                    # result 컬럼에 색상 적용 (카키)
                    apply_color(worksheet, row=cell.row, column=cell.column, color='CC9933')

                    # 호스트 컬럼에 색상 적용 (카키)
                    ip_column_index = result.columns.get_loc(host_column) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='CC9933')
                    
                    # IP 주소 컬럼에 색상 적용 (카키)
                    ip_column_index = result.columns.get_loc(ip_column[0]) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='CC9933')
                
                elif 'No IP Address' in cell.value:
                    # 'No IP Address' 행에 대해서만 분홍색 적용
                    for column in worksheet.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=worksheet.max_column):
                        for cell in column:
                            apply_color(worksheet, row=cell.row, column=cell.column, color='F6D8CE')

                    # result 컬럼에 색상 적용 (더 진한 분홍색)
                    apply_color(worksheet, row=cell.row, column=cell.column, color='FF8888')

                    # IP 주소 컬럼에 색상 적용 (더 진한 분홍색)
                    ip_column_index = result.columns.get_loc(ip_column[0]) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='FF8888')
                    
                    
                elif 'Invalid IP Address' in cell.value:
                    # 'Invalid IP Address' 행에 대해서만 주황색 적용
                    for column in worksheet.iter_cols(min_row=cell.row, max_row=cell.row, min_col=1, max_col=worksheet.max_column):
                        for cell in column:
                            apply_color(worksheet, row=cell.row, column=cell.column, color='ECCDB0')

                    # result 컬럼에 색상 적용 (더 진한 주황색)
                    apply_color(worksheet, row=cell.row, column=cell.column, color='E26B0A')

                    # IP 주소 컬럼에 색상 적용 (더 진한 주황색)
                    ip_column_index = result.columns.get_loc(ip_column[0]) + 1
                    apply_color(worksheet, row=cell.row, column=ip_column_index, color='E26B0A')
                
                else:
                    next

    # 컬럼 스타일 정의
    style = NamedStyle(name='Header')
    style.alignment = Alignment(horizontal='center', vertical='center')
    style.font = Font(name='맑은 고딕', bold=True)
    # 컬럼 스타일 적용
    for cell in worksheet[1]:
        cell.style = style
        apply_color(worksheet, row=cell.row, column=cell.column, color='D3D3D3')  # 연회색 적용

    apply_borders(worksheet, start_row=1, start_col=1, end_row=worksheet.max_row, end_col=worksheet.max_column)

    # 열의 너비 자동 조절
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        # 최대 길이 계산
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # 글자 수에 따라 조절할 너비 계산
        worksheet.column_dimensions[cell.column_letter].width = adjusted_width

    # 셀 정렬 (가운데 정렬)
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 파일 저장
    workbook.save(output_file)
    # print("결과를 저장한 파일:", output_file)

def apply_borders(worksheet, start_row, start_col, end_row, end_col):
    border_style = Side(border_style="thin", color="000000")
    border = Border(top=border_style, right=border_style, bottom=border_style, left=border_style)

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.border = border

def apply_color(worksheet, row, column, color):
    cell = worksheet.cell(row=row, column=column)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def get_all_ip_addresses(cidr):        #cidr 형식을 배열로 푸는 과정 192.168.0.1/24 -> [192.168.0.1, 192.168.0.2 ... 192.168.0.254]
    ip_network = ipaddress.ip_network(cidr)
    all_ip_addresses = [str(ip) for ip in ip_network.hosts()]
    return all_ip_addresses

def get_network_cidr(interface):  #네트워크 정보 가져오기 ip,서브넷 등등등..
    addresses = psutil.net_if_addrs()
    if interface in addresses:
        for addr in addresses[interface]:
            if addr.family == socket.AF_INET:
                return ipaddress.ip_interface((addr.address, addr.netmask)).network
    return None

def GetHostByAddress(ip_address):  #호스트네임 검색
    try:
        hostname = socket.gethostbyaddr(ip_address)[0]
        return hostname
    except socket.herror:
        return "Unknown"

def worker(queue, results):
    while True:
        ip = queue.get()
        if ip is None:
            break
        icmp_packet = IP(dst=ip)/ICMP()
        response = sr1(icmp_packet, timeout=1, verbose=False)
        if response:
            hostname = GetHostByAddress(ip)
            results[ip] = hostname
        queue.task_done()

def get_network_cidr_mapping():
    interfaces = psutil.net_if_stats().keys()
    mapping = {}
    for interface in interfaces:
        cidr = get_network_cidr(interface)
        if cidr is not None:
            mapping[interface] = str(cidr)
    return mapping

def test_connection(text):
    try:
        text = text.split(":")
        cidr_ipadress = text[1].strip()
        ip_addresses = get_all_ip_addresses(cidr_ipadress)
        num_threads = 50

        queue = Queue()
        results = {}
        threads = []
        ip_and_hostname = {}
        for _ in range(num_threads):
            t = threading.Thread(target=worker, args=(queue, results))
            t.start()
            threads.append(t)

        for ip in ip_addresses:
            queue.put(ip)

        queue.join()

        for _ in range(num_threads):
            queue.put(None)
        for t in threads:
            t.join()

        for ip, hostname in results.items():
            ip_and_hostname[ip] = hostname
    except:
        return False
    else:
        return ip_and_hostname

def tab_write_data(self, result):
    num_rows, num_cols = result.shape
    self.setRowCount(num_rows)
    self.setColumnCount(num_cols)
    
    # self.setColumnCount(len(result.columns))
    # i=0
    # for col in range(num_cols):
    #     header_item = QTableWidgetItem(result.columns[i])
    #     self.setHorizontalHeaderItem(col, header_item)
    #     i += 1
    
    for i in range(num_rows):
        for j in range(num_cols):
            item = QTableWidgetItem(str(result.iat[i, j]))
            self.setItem(i, j, item)

            if j == 0:
                    item.setTextAlignment(Qt.AlignCenter)

            if j == 1 and '대장o, 스캔x' in result['result'].iloc[i]:
                item.setBackground(QColor(248, 255, 46))
            elif j == 1 and '대장x 스캔o' in result['result'].iloc[i]:
                item.setBackground(QColor(254, 46, 46))
            elif j == 1 and '스캔불가' in result['result'].iloc[i]:
                item.setBackground(QColor(168, 168, 168))
            


            

# QmainWindow -> mywindow 상속 (PyQt에서 기본적인 창 생성하기 위한 클래스)
class MyWindow(QMainWindow):
    def __init__(self): #클래스 초기화, self매개변수는 객체 자체를 나타낸다.
        super().__init__()
        self.initUI() #gui초기화, 사용자인터이스요소 ?

    def initUI(self):
        self.setWindowIcon(QIcon(''))
        self.setWindowTitle('IP 관리대장 점검 시스템')
        self.resize(1000, 600)
        self.center()
                
        # self.progress_bar = QProgressBar(self)
        # self.progress_bar.setGeometry(70,305,560,40)
        # self.progress_bar.setStyleSheet("border: 5px solid; background-color: white")

        self.font = QLabel('@Copyrights by TMI Potato',self)
        self.font.setGeometry(875,555,230,50)
        font = self.font.font()
        font.setPointSize(6)
        font.setWeight(QFont.Bold)
        self.font.setFont(font)

        menubar = self.menuBar()
        menu = menubar.addMenu('도움말')
        help = QAction('도움말 보기', self)
        help.triggered.connect(self.help_load)
        help.setShortcut('Ctrl+h')
        menu.addAction(help)  

        # 한줄짜리 글자를 입력받을 수 있는 입력위젯 QLineEdit (저장위치)
        self.push_line_edit = QLineEdit(self)
        self.push_line_edit.setReadOnly(True)
        self.push_line_edit.setGeometry(15, 30, 350, 30)
        self.push_line_edit.setFont(QFont('나눔고딕',10))

        self.pushButton = QPushButton('△불러오기', self)
        self.pushButton.clicked.connect(self.pushButtonClicked)
        self.pushButton.setGeometry(370, 30, 80, 30)   
        self.pushButton.setFont(QFont('나눔고딕',10))

        #스캔대역 
        self.cb = QComboBox(self)
        for interface, cidr in get_network_cidr_mapping().items():
            self.cb.addItem(f" {interface}: {cidr}")
        self.cb.setGeometry(460, 30, 350, 30)
        self.cb.setFont(QFont('나눔고딕',10))

        self.scanbutton = QPushButton('SCAN', self)
        self.scanbutton.setShortcut("s") #단축키 
        self.scanbutton.setGeometry(815, 30, 80, 30)
        self.scanbutton.setFont(QFont('나눔고딕',10))
        self.scanbutton.clicked.connect(self.scanbuttonClicked)

        self.savebutton = QPushButton('▽CSV저장', self)
        self.savebutton.clicked.connect(self.savebuttonClicked)
        self.savebutton.setGeometry(905, 30, 80, 30)
        self.savebutton.setFont(QFont('나눔고딕',10))

        # self.pushButton.setStyleSheet("border: 1px solid; background-color: white")
        # self.savebutton.setStyleSheet("border: 1px solid; background-color: white")
        # self.scanbutton.setStyleSheet("border: 1px solid; background-color: white")
        # self.cb.setStyleSheet("border: 1px solid; background-color: white")
        self.push_line_edit.setStyleSheet("border: 1px solid lightgray; background-color: white")
        self.cb.setStyleSheet("border: 1px solid lightgray; background-color: white")

        self.showTabs()

    def showTabs(self):

        tab1_colum = 5
        tab1_row = 0
        columns_names = ['번호','#','IP 주소', '호스트 이름', '진단결과']

        # 첫번째 탭 
        tab1 = QWidget()
        self.table1 = QTableWidget()
        self.table1.setColumnCount(tab1_colum)
        self.table1.setRowCount(tab1_row)
        self.table1.setColumnWidth(0, 45)
        self.table1.setColumnWidth(1, 5)
        self.table1.setColumnWidth(2, 200)
        self.table1.setColumnWidth(3, 200)
        self.table1.setColumnWidth(4, 470)
        self.table1.verticalHeader().setVisible(False)
        self.tabs = QTabWidget(self)
        self.tabs.addTab(tab1, '전체')

        
        
        self.table1.setHorizontalHeaderLabels(columns_names)

        layout1 = QVBoxLayout(tab1)
        layout1.addWidget(self.table1)

        # 두번째 탭
        tab2 = QWidget()
        self.table2 = QTableWidget()
        self.table2.setColumnCount(tab1_colum)
        self.table2.setRowCount(tab1_row)
        self.table2.setColumnWidth(0, 45)
        self.table2.setColumnWidth(1, 5)
        self.table2.setColumnWidth(2, 200)
        self.table2.setColumnWidth(3, 200)
        self.table2.setColumnWidth(4, 470)
        self.table2.verticalHeader().setVisible(False)
        self.tabs.addTab(tab2, '대장o, 스캔x')

        self.tabs.setGeometry(15,70,975,500)

        self.table2.setHorizontalHeaderLabels(columns_names)

        vbox = QVBoxLayout(tab2)
        vbox.addWidget(self.table2)
        
        # 세번째 탭
        tab3 = QWidget()
        self.table3 = QTableWidget()
        self.table3.setColumnCount(tab1_colum)
        self.table3.setRowCount(tab1_row)
        self.table3.setColumnWidth(0, 45)
        self.table3.setColumnWidth(1, 5)
        self.table3.setColumnWidth(2, 200)
        self.table3.setColumnWidth(3, 200)
        self.table3.setColumnWidth(4, 470)
        self.table3.verticalHeader().setVisible(False)
        self.tabs.addTab(tab3, '대장x 스캔o')

        self.tabs.setGeometry(15,70,975,500)

        self.table3.setHorizontalHeaderLabels(columns_names)

        vbox = QVBoxLayout(tab3)
        vbox.addWidget(self.table3)

        # 네번째 탭
        tab4 = QWidget()
        self.table4 = QTableWidget()
        self.table4.setColumnCount(tab1_colum)
        self.table4.setRowCount(tab1_row)
        self.table4.setColumnWidth(0, 45)
        self.table4.setColumnWidth(1, 5)
        self.table4.setColumnWidth(2, 200)
        self.table4.setColumnWidth(3, 200)
        self.table4.setColumnWidth(4, 470)
        self.table4.verticalHeader().setVisible(False)

        self.tabs.addTab(tab4, '스캔불가')

        self.tabs.setGeometry(15,70,975,500)

        self.table4.setHorizontalHeaderLabels(columns_names)

        vbox = QVBoxLayout(tab4)
        vbox.addWidget(self.table4)
        
        # 다섯번째 탭
        tab5 = QWidget()
        self.table5 = QTableWidget()
        self.table5.setColumnCount(tab1_colum)
        self.table5.setRowCount(tab1_row)
        self.table5.setColumnWidth(0, 45)
        self.table5.setColumnWidth(1, 5)
        self.table5.setColumnWidth(2, 200)
        self.table5.setColumnWidth(3, 200)
        self.table5.setColumnWidth(4, 470)
        self.table5.verticalHeader().setVisible(False)

        self.tabs.addTab(tab5, '호스트불일치')

        self.table5.setHorizontalHeaderLabels(columns_names)

        self.tabs.setGeometry(15,70,975,500)

        vbox = QVBoxLayout(tab5)
        vbox.addWidget(self.table5)
        
    def help_load(self):
        url = QUrl('https://github.com/Phqasue/Potato-IP-Scanner')
        QDesktopServices.openUrl(url)
        
    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()

        qr.moveCenter(cp)
        self.move(qr.topLeft())

    def pushButtonClicked(self):
        fname = QFileDialog.getOpenFileName(self)
        self.push_line_edit.setText(fname[0])

    def saveFileDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog

        selected_dir = QFileDialog.getExistingDirectory(self, '저장 폴더 선택', options=options)
        return selected_dir

    def scanbuttonClicked(self):
        push_file = self.push_line_edit.text()
        
        if push_file == '':
            QMessageBox.warning(self, '경고', 'IP 관리 대장을 불러와주세요.')
            return
        elif os.path.splitext(push_file)[1] != '.xlsx':
            QMessageBox.warning(self, '경고', '올바르지 않은 파일 형식입니다.')
            return
        if not os.path.exists(push_file):
            QMessageBox.warning(self, '경고', '불러온 파일이 존재하지 않습니다.')
            return
        
        QMessageBox.information(self, '알림', '스캔이 정상적으로 시작이 되었습니다.')
        #self.scanbutton.setEnabled(False) #scan버튼 비활성화
        # self.progress_bar.setValue(0)
        # self.progress_bar.setFormat("%p%")  # 진행 상황 표시 형식 설정
        # self.progress_bar.setAlignment(Qt.AlignCenter)  # 진행 상황 텍스트 가운데 정렬
        # self.thread = progress()
        # self.thread.update_progress.connect(self.update_progress_bar)
        # self.thread.finished.connect(self.progress_finished)
        # self.thread.start()

        ipaddress = self.cb.currentText() #combobox 'cb' -> 현재 선택된 ip주소
        scan_result = test_connection(ipaddress) #스캔결과

        if scan_result == False:
            QMessageBox.information(self, '알림', '스캔중 오류가 발생했습니다.\n 네트워크 대역을 다시 확인해주세요.')
            return
        else:
            next
        
        result = Verification(push_file, scan_result)

        

        if not result.empty:
            QMessageBox.information(self, '알림', '스캔이 완료되었습니다.') 
            tab_write_data(self.table1, result)
            # setRowBackground(self.table1, result.shape[0])


            tab_write_data(self.table2, result[result['result'].str.contains('대장o, 스캔x')])
            tab_write_data(self.table3, result[result['result'].str.contains('대장x 스캔o')])
            tab_write_data(self.table4, result[result['result'].str.contains('스캔불가')])
            tab_write_data(self.table5, result[result['result'].str.contains('호스트')])

            for i in range(result.shape[0]):
                if result['result'].str.contains('대장o, 스캔x').iloc[i]:
                    item = self.table1.item(i, 1)
                    item.setBackground(QColor(248, 255, 46))  
                
            for i in range(result.shape[0]):
                if result['result'].str.contains('대장x 스캔o').iloc[i]:
                    item = self.table1.item(i, 1)
                    item.setBackground(QColor(254, 46, 46))    
            
            for i in range(result.shape[0]):
                if result['result'].str.contains('스캔불가').iloc[i]:
                    item = self.table1.item(i, 1)
                    item.setBackground(QColor(168, 168, 168))

            for i in range(result.shape[0]):
                if result['result'].iloc[i] == '' or pd.isna(result['result'].iloc[i]):
                    item = self.table1.item(i, 1)
                    item.setBackground(QColor(0, 255, 64))

        
    def savebuttonClicked(self):
        push_file = self.push_line_edit.text()
        save_text = self.saveFileDialog()

        file_name_without_ext, file_ext = os.path.splitext(os.path.basename(push_file)) #관리대장 파일 이름/확장자 추출
        new_file_name = file_name_without_ext + "_Result" + file_ext #저장할 파일명
        new_file_dir = save_text + '/' + new_file_name #저장할 파일경로
        

        # QTableWidget에서 데이터 가져와서 Pandas DataFrame으로 변환
        num_rows = self.table1.rowCount()
        num_cols = self.table1.columnCount()
        data = []
        for row in range(num_rows):
            row_data = []
            for col in range(num_cols):
                item = self.table1.item(row, col)
                if item is not None:
                    row_data.append(item.text())
                else:
                    row_data.append('')
            data.append(row_data)
        
        df = pd.DataFrame(data)



        # Pandas DataFrame을 엑셀 파일로 저장
        df.to_excel(new_file_dir, index=False, header=True)
        
    
        
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    app.exec_()